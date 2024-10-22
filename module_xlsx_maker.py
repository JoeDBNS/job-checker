# Column filters
# https://openpyxl.readthedocs.io/en/stable/filters.html#using-filters-and-sorts



import os, datetime as dt
from openpyxl import Workbook
from openpyxl.worksheet.filters import (
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    FilterColumn,
    Filters,
)
from openpyxl.styles import (
    Alignment,
    PatternFill
)



def Setup():
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet('data')
    ws = wb["data"]
    return wb, ws


def SaveFile(wb, name):
    # YYYYMMDD
    timestamp_today = str(dt.date.today()).replace('-', '')

    reports_path = './_reports/' + timestamp_today

    if not os.path.exists(reports_path):
        os.makedirs(reports_path)

    wb.save(reports_path + '/' + name + '.xlsx')

    return (reports_path + '/' + name + '.xlsx')


def SetColumnColors(ws, column_colors):
    for idx, row in enumerate(ws.rows):
        for col in row:
            ws[col.coordinate].fill = PatternFill(start_color=column_colors[idx], end_color=column_colors[idx], fill_type='solid')


def SetColumnSize(ws, column_sizes):
    for idx, col in enumerate(ws.columns):
        max_length = 7
        column_letter = col[0].column_letter

        if (column_sizes[idx] == 0):
            for cell in col:
                cell.alignment = Alignment(horizontal='left')

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as error:
                    pass

            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        else:
            ws.column_dimensions[column_letter].width = column_sizes[idx]


def SetFindingsHyperlinks(ws):
    for col in ws.columns:
        column_letter = col[0].column_letter

        # hard coding column with hyperlinks
        if(column_letter == 'K'):
            for cell in col[1:]:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def BuildXlsxFile(name, data, column_colors, column_sizes):
    wb, ws = Setup()

    for row in data:
        ws.append(row)

    SetColumnSize(ws, column_sizes)

    SetFindingsHyperlinks(ws)
    SetColumnColors(ws, column_colors)
    SetColumnSize(ws, column_sizes)

    return SaveFile(wb, name)