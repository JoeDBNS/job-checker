# Column filters
# https://openpyxl.readthedocs.io/en/stable/filters.html#using-filters-and-sorts

# Visuals - Charts / Graphs
# https://openpyxl.readthedocs.io/en/stable/charts/introduction.html



import os, datetime as dt
from openpyxl import Workbook
from openpyxl.worksheet.filters import (
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    FilterColumn,
    Filters,
)
from openpyxl.styles import (
    Alignment,
    PatternFill
)



def CreateWorkbook(sheet_content):
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in sheet_content:
        wb.create_sheet(sheet['name'])

    return wb


def SetColumnHyperlinks(wb, sheet_content):
    link_columns = []
    sheet_has_headers = sheet_content['config']['general']['add_headers']

    for i, col in enumerate(sheet_content['config']['columns']):
        if (col['is_link'] == True):
            link_columns.append(i)

    if (len(link_columns) > 0):
        for i, row in enumerate(wb[sheet_content['name']].rows):
            # Skip headers row if headers are added
            if (i == 0 and sheet_has_headers):
                pass
            else:
                for id in link_columns:
                    row[id].hyperlink = row[id].value
                    row[id].style = "Hyperlink"


def SetColumnSizes(wb, sheet_content):
    for i, col in enumerate(wb[sheet_content['name']].columns):
        min_length = 7
        column_letter = col[0].column_letter

        if (len(sheet_content['config']['columns']) > 0):
            # Later Task: break out cell.alignment into own function
            if (sheet_content['config']['columns'][i]['size'] == 0):
                for cell in col:
                    cell.alignment = Alignment(horizontal='left')

                    # Column width is defined by character count
                    try:
                        if len(str(cell.value)) > min_length:
                            min_length = len(cell.value)
                    except Exception as error:
                        pass

                adjusted_width = (min_length + 2)
                wb[sheet_content['name']].column_dimensions[column_letter].width = adjusted_width

            else:
                for cell in col:
                    cell.alignment = Alignment(horizontal='left')

                wb[sheet_content['name']].column_dimensions[column_letter].width = sheet_content['config']['columns'][i]['size']


def SetColumnColors(wb, sheet_content):
    row_colors_ids = []
    row_colors = {}

    for color_def in sheet_content['config']['rows']['colors']:
        row_colors_ids.append(color_def['row_num'])
        row_colors[color_def['row_num']] = color_def['color_hex']

    for i, row in enumerate(wb[sheet_content['name']].rows):

        if (i in row_colors_ids):
            for col in row:
                wb[sheet_content['name']][col.coordinate].fill = PatternFill(start_color=row_colors[i], end_color=row_colors[i], fill_type='solid')


def ConfigureWorksheet(wb, sheet_content):
    SetColumnHyperlinks(wb, sheet_content)
    SetColumnSizes(wb, sheet_content)
    SetColumnColors(wb, sheet_content)


def SaveFile(wb, name):
    # YYYYMMDD
    timestamp_today = str(dt.date.today()).replace('-', '')

    reports_path = './_reports/' + timestamp_today

    if not os.path.exists(reports_path):
        os.makedirs(reports_path)

    wb.save(reports_path + '/' + name + '.xlsx')

    return (reports_path + '/' + name + '.xlsx')


def BuildXlsxFile(wb_content):
    wb = CreateWorkbook(wb_content['sheets'])

    for sheet_content in wb_content['sheets']:
        ws = wb[sheet_content['name']]

        if (sheet_content['config']['general']['add_headers'] == True):
            header_data = []
            for column_config in sheet_content['config']['columns']:
                header_data.append(column_config['label'])
            ws.append(header_data)

        for row in sheet_content['data']:
            ws.append(row)

        ConfigureWorksheet(wb, sheet_content)

    return SaveFile(wb, wb_content['name'])




# incoming object definition:
# wb_content['sheets']['name'] must be unique.
# wb_content['sheets']['data'] must ultimately be a list of lists.
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# wb_content = {
#     'name': '',
#     'sheets': [
#         {
#             'name': '',
#             'config': {
#                 'general': {
#                     'add_headers': True
#                 },
#                 'columns': [
#                     {
#                         'label': '',
#                         'size': 0,
#                         'filter': False,
#                         'is_link': False
#                     }
#                 ],
#                 'rows': {
#                     'colors': [
#                         {
#                             'row_num': 0,
#                             'color_hex': ''
#                         }
#                     ]
#                 }
#             },
#             'data': []
#         }
#     ]
# }